import openpyxl
from pathlib import Path

excelFile = openpyxl.Workbook()
print(excelFile.sheetnames)

excelSheet = excelFile.active
excelSheet.title = "firstSheet"

excelFile.create_sheet()
excelFile.create_sheet(index=1, title= "secondSheet")

sheet = excelFile['secondSheet']
sheet['A1'] = 'Hello, world!'
print(sheet['A1'].value)

names = ['hadi','yara','sara','anas']
firstSheet = excelFile['firstSheet']
for i in range(1,len(names)+1):
    firstSheet.cell(row=i,column=3).value =names[i-1]


#SAVE EXCEL FILE
excelFile.save(filename = 'newExcel.xlsx')
